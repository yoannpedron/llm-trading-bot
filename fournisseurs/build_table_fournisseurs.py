# -*- coding: utf-8 -*-
"""
Consolide l'annuaire fournisseurs Audika en une table simple a 4 colonnes.

Sources :
  - Annuaire_FRN.xlsx  (onglets "annuaire FRN contrats" et "annuaire FRN espaces verts")
  - Audika_Table_correspondance_centres.xlsx (contexte : document Eurofeu, pas d'annuaire FRN)

Sortie : Table_fournisseurs_consolidee.xlsx
  Onglet 1 "Fournisseurs"              -> Fournisseur | Domaine d'activite | Coordonnees telephoniques | Personne a contacter (telephone)
  Onglet 2 "Detail contacts"           -> une ligne par contact, toutes les colonnes sources conservees
  Onglet 3 "Sources & a verifier"      -> provenance + anomalies relevees pendant la consolidation
"""
import re
import sys
import unicodedata
from collections import OrderedDict

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SRC_FRN = sys.argv[1]
SRC_CENTRES = sys.argv[2]
OUT = sys.argv[3]

# --------------------------------------------------------------------------- outils

def clean(v):
    """Normalise une cellule en texte propre (espaces/retours ligne/tabulations)."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    s = str(v)
    s = s.replace("\r", " ").replace("\n", " ").replace("\t", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s


def fmt_phone(v):
    """Renvoie (numero_formate, note). Les numeros stockes en nombre ont perdu leur 0 initial."""
    s = clean(v)
    if not s:
        return "", ""
    digits = re.sub(r"\D", "", s)
    if len(digits) == 9 and not digits.startswith("0"):
        digits = "0" + digits
    if len(digits) == 10 and digits.startswith("0"):
        num = " ".join(digits[i:i + 2] for i in range(0, 10, 2))
        reste = re.sub(r"[\d\s./-]+", " ", s).strip(" ,;()")
        return num, reste
    if len(digits) in (11, 12) and digits.startswith("33"):
        digits = "0" + digits[2:]
        return " ".join(digits[i:i + 2] for i in range(0, 10, 2)), ""
    return "", s  # pas un numero exploitable : on conserve le texte tel quel


def norm_key(name):
    s = unicodedata.normalize("NFKD", clean(name))
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"[^A-Za-z0-9]+", " ", s).upper().strip()
    return s


# Libelles generiques (services, agences, standards) : ils alimentent la colonne
# "Coordonnees telephoniques" et non "Personne a contacter".
GENERIC_STARTS = (
    "standard", "sav", "service", "agence", "plateforme", "pole", "info",
    "demande", "comptabilite", "generique", "maintenance", "planning",
    "gestion", "contact", "depannage", "boite", "courriel", "departement",
    "generaliste", "assistante d", "travaux", "siege", "vpi", "adv",
    "accueil", "secretariat",
)


# Libelles generiques que le prefixe ne permet pas de detecter (agence nommee par sa ville).
MANUAL_GENERIC = {"saint denis"}


def is_generic(label):
    key = norm_key(label).lower()
    if not key:
        return True
    if key in MANUAL_GENERIC:
        return True
    return any(key.startswith(g) for g in GENERIC_STARTS)


def add_domain(s, dom):
    """Ajoute un domaine en evitant les doublons et les libelles inclus l'un dans l'autre."""
    dom = clean(dom)
    if not dom:
        return
    k = norm_key(dom)
    for i, existant in enumerate(s["domaines"]):
        ke = norm_key(existant)
        if k == ke or k in ke:
            return
        if ke in k:                      # le nouveau libelle est plus complet
            s["domaines"][i] = dom
            return
    s["domaines"].append(dom)


# --------------------------------------------------------------------------- lecture

anomalies = []          # (fournisseur, constat, incidence)
suppliers = OrderedDict()   # cle normalisee -> dict


def get_supplier(name, domain="", compte="", region="", centre=""):
    key = norm_key(name)
    s = suppliers.get(key)
    if s is None:
        s = {
            "nom": clean(name),
            "domaines": [],
            "comptes": [],
            "regions": [],
            "centres": [],
            "notes": [],
            "generiques": [],   # (libelle, [numeros], note)
            "personnes": [],    # (nom, fonction, [numeros])
            "contacts": [],     # lignes detail
            "occurrences": [],  # (onglet, domaine) de chaque entree source
        }
        suppliers[key] = s
    add_domain(s, domain)
    for champ, val in (("comptes", compte), ("regions", region), ("centres", centre)):
        val = clean(val)
        if val and val not in s[champ]:
            s[champ].append(val)
    return s


def add_contact(s, contact, fonction, tel, portable, courriel, commentaire, source,
                region="", centre="", header_row=False):
    contact, fonction = clean(contact), clean(fonction)
    nums, notes = [], []
    for raw in (tel, portable):
        num, note = fmt_phone(raw)
        if num and num not in nums:
            nums.append(num)
        if note:
            notes.append(note)

    s["contacts"].append([
        s["nom"], " ; ".join(s["domaines"]), " / ".join(s["comptes"]),
        clean(region), clean(centre), contact, fonction,
        " / ".join(nums) if nums else "", clean(courriel),
        " ; ".join([clean(commentaire)] + notes).strip(" ;"), source,
    ])

    if not nums and not notes:
        return
    label = contact or fonction
    if is_generic(label):
        if not label:
            # numero porte par la ligne d'en-tete du fournisseur -> numero principal ;
            # sinon ligne isolee sans libelle -> a verifier.
            label = "Standard" if header_row else "N° sans libellé (à vérifier)"
            if not header_row:
                anomalies.append((s["nom"],
                                  "Numéro présent dans la source sans libellé ni nom associé : "
                                  + ", ".join(nums) + ".",
                                  "Rattachement du numéro à confirmer."))
        s["generiques"].append((label, nums, "; ".join(notes)))
    else:
        s["personnes"].append((contact, fonction, nums + notes))


# --- onglet "annuaire FRN contrats" ------------------------------------------------
wb = openpyxl.load_workbook(SRC_FRN, data_only=True)
ws = wb["annuaire FRN contrats"]
courant = None
for row in ws.iter_rows(min_row=2, values_only=True):
    ent, compte, region, centre, domaine, contact, fonction, tel, port, mail, comm = \
        (list(row) + [None] * 11)[:11]
    if not any(clean(c) for c in row):
        continue

    entete = bool(clean(ent) and any(clean(x) for x in (compte, region, centre, domaine)))
    if entete:
        courant = get_supplier(ent, domaine, compte, region, centre)
        courant["occurrences"].append(("annuaire FRN contrats", clean(domaine) or "(domaine vide)"))
    elif clean(ent):
        # ligne de continuation : suite du nom ou reference de compte client
        if courant is not None:
            courant["notes"].append(clean(ent))
    if courant is None:
        continue

    for champ, val in (("regions", region), ("centres", centre)):
        val = clean(val)
        if val and val not in courant[champ]:
            courant[champ].append(val)
    add_domain(courant, domaine)

    if any(clean(x) for x in (contact, fonction, tel, port, mail, comm)):
        add_contact(courant, contact, fonction, tel, port, mail, comm,
                    "annuaire FRN contrats", region, centre, header_row=entete)

# --- onglet "annuaire FRN espaces verts" -------------------------------------------
ws = wb["annuaire FRN espaces verts"]
courant = None
for row in ws.iter_rows(min_row=2, values_only=True):
    ent, compte, centre, contact, fonction, tel, port, mail, comm = \
        (list(row) + [None] * 9)[:9]
    if not any(clean(c) for c in row):
        continue
    entete = bool(clean(ent))
    if entete:
        courant = get_supplier(ent, "ESPACES VERTS / PAYSAGISTE", compte, "", centre)
        courant["occurrences"].append(("annuaire FRN espaces verts", "ESPACES VERTS / PAYSAGISTE"))
    if courant is None:
        continue
    if any(clean(x) for x in (contact, fonction, tel, port, mail, comm)):
        add_contact(courant, contact, fonction, tel, port, mail, comm,
                    "annuaire FRN espaces verts", "", centre, header_row=entete)

# ------------------------------------------------------- doublons fusionnes detectes
for s_ in suppliers.values():
    if len(s_["occurrences"]) > 1:
        det = " + ".join(f"« {o} » ({d})" for o, d in s_["occurrences"])
        anomalies.append((s_["nom"],
                          f"Saisi {len(s_['occurrences'])} fois dans les sources : {det}.",
                          "Lignes fusionnées en une seule ; domaines cumulés."))

# --------------------------------------------------------------- corrections ciblees
DOMAINE_MANQUANT = {"TECH 9 ENERGIE": "CLIMATICIEN"}
for nom, dom in DOMAINE_MANQUANT.items():
    s = suppliers.get(norm_key(nom))
    if s and not s["domaines"]:
        add_domain(s, dom)
        anomalies.append((s["nom"],
                          "Colonne DOMAINE D'ACTIVITÉ vide dans la source ; « Climaticien » "
                          "avait été saisi dans la colonne Centre.",
                          "Domaine rétabli à CLIMATICIEN dans la table consolidée."))

ANOMALIES_MANUELLES = [
    ("GRAF SERVICES PLUS",
     "Les trois contacts « @gestivert.fr » (C. Foissin, A. Mauclair, C. Dos Santos) suivent "
     "GRAF SERVICES PLUS alors que le premier contact est « @stihle.fr » : nom de société "
     "probablement manquant dans la source.",
     "Rattachés à GRAF SERVICES PLUS par défaut ; à confirmer, GESTIVERT est peut-être un "
     "fournisseur distinct."),
    ("CLIMSERV (et CLIM' TECH travaux install)",
     "« Groupe GES’ TECH » figure en colonne ENTREPRISE sur la ligne suivante : suite du nom, "
     "pas un nouveau fournisseur.",
     "Conservé comme information de groupe, aucun fournisseur créé."),
    ("EUROFEU",
     "Comptes clients saisis en colonne ENTREPRISE : C660995 (SOGECA) et C660964 (AUDIKA ALPES).",
     "Repris en information ; un seul fournisseur EUROFEU dans la table."),
    ("Toutes lignes",
     "De nombreux numéros sont stockés au format nombre dans la source : le 0 initial est perdu "
     "(ex. 299602704).",
     "Zéro rétabli et numéros remis au format 02 99 60 27 04."),
]
anomalies.extend(ANOMALIES_MANUELLES)

# --------------------------------------------------------------------------- ecriture
ARIAL = "Arial"
F_TITRE = Font(name=ARIAL, size=14, bold=True, color="1F3864")
F_SOUS = Font(name=ARIAL, size=9, italic=True, color="595959")
F_HEAD = Font(name=ARIAL, size=10, bold=True, color="FFFFFF")
F_CELL = Font(name=ARIAL, size=10)
F_NOM = Font(name=ARIAL, size=10, bold=True)
FILL_HEAD = PatternFill("solid", fgColor="1F3864")
FILL_ALT = PatternFill("solid", fgColor="F2F5FA")
FILL_WARN = PatternFill("solid", fgColor="FFF2CC")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TOP = Alignment(vertical="top", wrap_text=True)
TOPL = Alignment(vertical="top", wrap_text=True, horizontal="left")

out = openpyxl.Workbook()

# ---- Onglet 1 : la table demandee -------------------------------------------------
s1 = out.active
s1.title = "Fournisseurs"
s1["A1"] = "Table fournisseurs consolidee"
s1["A1"].font = F_TITRE
s1["A2"] = ("Sources : Annuaire_FRN.xlsx (onglets « annuaire FRN contrats » et « annuaire FRN "
            "espaces verts »). Une ligne par fournisseur ; doublons fusionnés. "
            "Détail complet des contacts en onglet « Détail contacts ».")
s1["A2"].font = F_SOUS
s1.merge_cells("A1:D1")
s1.merge_cells("A2:D2")
s1.row_dimensions[2].height = 26
s1["A2"].alignment = Alignment(vertical="center", wrap_text=True)

HEAD1 = ["Fournisseur", "Domaine d’activité", "Coordonnées téléphoniques",
         "Personne à contacter (téléphone)"]
for c, h in enumerate(HEAD1, 1):
    cell = s1.cell(row=4, column=c, value=h)
    cell.font, cell.fill, cell.border = F_HEAD, FILL_HEAD, BORDER
    cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
s1.row_dimensions[4].height = 22

sans_tel = []
r = 5
for s in sorted(suppliers.values(), key=lambda x: norm_key(x["nom"])):
    gen_txt = []
    for label, nums, note in s["generiques"]:
        if nums:
            txt = f"{label} : " + " / ".join(nums)
        elif note:
            txt = f"{label} : {note}"
        else:
            continue
        if note and nums:
            txt += f" ({note})"
        if txt not in gen_txt:
            gen_txt.append(txt)

    fusion = OrderedDict()
    for nom, fonction, nums in s["personnes"]:
        if not nums:
            continue
        cle = (norm_key(nom), tuple(nums))
        if cle in fusion:
            if fonction and fonction not in fusion[cle][1]:
                fusion[cle][1].append(fonction)
        else:
            fusion[cle] = (nom, [fonction] if fonction else [], nums)
    pers_txt = []
    for nom, fonctions, nums in fusion.values():
        txt = nom
        if fonctions:
            txt += " (" + " / ".join(fonctions) + ")"
        txt += " : " + " / ".join(nums)
        if txt not in pers_txt:
            pers_txt.append(txt)

    if not gen_txt and not pers_txt:
        sans_tel.append(s["nom"])

    vals = [
        s["nom"],
        " ; ".join(s["domaines"]) or "—",
        "\n".join(gen_txt) or "—",
        "\n".join(pers_txt) or "—",
    ]
    for c, v in enumerate(vals, 1):
        cell = s1.cell(row=r, column=c, value=v)
        cell.font = F_NOM if c == 1 else F_CELL
        cell.alignment = TOPL
        cell.border = BORDER
        if r % 2 == 1:
            cell.fill = FILL_ALT
    r += 1

last1 = r - 1
WIDTHS1 = (30, 26, 34, 62)


def hauteur(valeurs, largeurs, mini=15, maxi=409):
    """Hauteur de ligne approchee pour du texte renvoye a la ligne."""
    lignes = 1
    for v, w in zip(valeurs, largeurs):
        n = 0
        for morceau in str(v).split("\n"):
            n += max(1, -(-len(morceau) // max(8, int(w) - 2)))
        lignes = max(lignes, n)
    return min(maxi, max(mini, lignes * 12.9 + 3))


for rr in range(5, last1 + 1):
    s1.row_dimensions[rr].height = hauteur(
        [s1.cell(row=rr, column=c).value or "" for c in range(1, 5)], WIDTHS1)

for col, w in zip("ABCD", WIDTHS1):
    s1.column_dimensions[col].width = w
s1.freeze_panes = "A5"
s1.auto_filter.ref = f"A4:D{last1}"

# ---- Onglet 2 : detail ------------------------------------------------------------
s2 = out.create_sheet("Détail contacts")
HEAD2 = ["Fournisseur", "Domaine d’activité", "N° compte FRN Audika", "Région",
         "Centre(s)", "Contact", "Fonction", "Téléphone(s)", "Courriel",
         "Commentaire", "Source (onglet)"]
for c, h in enumerate(HEAD2, 1):
    cell = s2.cell(row=1, column=c, value=h)
    cell.font, cell.fill, cell.border = F_HEAD, FILL_HEAD, BORDER
    cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
s2.row_dimensions[1].height = 24

r = 2
for s in sorted(suppliers.values(), key=lambda x: norm_key(x["nom"])):
    for ligne in s["contacts"]:
        ligne = list(ligne)
        ligne[1] = " ; ".join(s["domaines"])
        for c, v in enumerate(ligne, 1):
            cell = s2.cell(row=r, column=c, value=v)
            cell.font = F_CELL
            cell.alignment = TOPL
            cell.border = BORDER
            if r % 2 == 0:
                cell.fill = FILL_ALT
        r += 1
last2 = r - 1
WIDTHS2 = (28, 24, 26, 12, 26, 28, 30, 24, 32, 40, 22)
for rr in range(2, last2 + 1):
    s2.row_dimensions[rr].height = hauteur(
        [s2.cell(row=rr, column=c).value or "" for c in range(1, 12)], WIDTHS2)
for col, w in zip("ABCDEFGHIJK", WIDTHS2):
    s2.column_dimensions[col].width = w
s2.freeze_panes = "A2"
s2.auto_filter.ref = f"A1:K{last2}"

# ---- Onglet 3 : sources & points a verifier ---------------------------------------
s3 = out.create_sheet("Sources & à vérifier")
s3["A1"] = "Provenance des données"
s3["A1"].font = F_TITRE
lignes_src = [
    ("Annuaire_FRN.xlsx — « annuaire FRN contrats »",
     "Source principale : fournisseurs sous contrat, domaine d’activité, contacts, téléphones."),
    ("Annuaire_FRN.xlsx — « annuaire FRN espaces verts »",
     "Fusionné dans la même table, domaine « ESPACES VERTS / PAYSAGISTE »."),
    ("Audika_Table_correspondance_centres.xlsx",
     "Table de correspondance des centres transmise à Eurofeu (632 centres, anomalies CP, "
     "intégrations 2026). Ne contient aucune coordonnée fournisseur : rien à reprendre ici. "
     "Eurofeu figure déjà dans la table (SÉCURITÉ INCENDIE)."),
    ("Fichier de Lucile (Teams)",
     "Non fourni avec cette demande : à intégrer dès réception, la structure est prête."),
]
r = 3
for a, b in lignes_src:
    s3.cell(row=r, column=1, value=a).font = F_NOM
    s3.cell(row=r, column=2, value=b).font = F_CELL
    s3.cell(row=r, column=1).alignment = TOPL
    s3.cell(row=r, column=2).alignment = TOPL
    r += 1

r += 1
s3.cell(row=r, column=1, value="Points à vérifier").font = F_TITRE
r += 2
for c, h in enumerate(["Fournisseur", "Constat", "Incidence / action"], 1):
    cell = s3.cell(row=r, column=c, value=h)
    cell.font, cell.fill, cell.border = F_HEAD, FILL_HEAD, BORDER
    cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
r += 1

sans_compte = sorted(x["nom"] for x in suppliers.values() if not x["comptes"])
if sans_compte:
    anomalies.append((", ".join(sans_compte),
                      "Aucun n° de compte fournisseur Audika renseigné dans la source.",
                      "À créer / à rapprocher de la base fournisseurs."))

if sans_tel:
    anomalies.append((", ".join(sorted(sans_tel)),
                      "Aucun numéro de téléphone dans la source.",
                      "Coordonnées à compléter (« — » dans la table)."))

for a, b, c in anomalies:
    for col, v in enumerate((a, b, c), 1):
        cell = s3.cell(row=r, column=col, value=v)
        cell.font = F_CELL
        cell.alignment = TOPL
        cell.border = BORDER
        cell.fill = FILL_WARN
    r += 1

WIDTHS3 = (44, 68, 46)
for rr in range(3, r):
    s3.row_dimensions[rr].height = hauteur(
        [s3.cell(row=rr, column=c).value or "" for c in range(1, 4)], WIDTHS3)
for col, w in zip("ABC", WIDTHS3):
    s3.column_dimensions[col].width = w

out.save(OUT)
print(f"OK -> {OUT}")
print(f"   fournisseurs : {len(suppliers)}")
print(f"   contacts     : {last2 - 1}")
print(f"   anomalies    : {len(anomalies)}")
